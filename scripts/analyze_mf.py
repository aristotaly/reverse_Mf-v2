"""Re-parse MacroFactor export with corrected column alignment.

The MF export has two INDEPENDENT arrays:
  - Date + Scale Weight: only logged days (~443 rows)
  - Trend Weight: continuous daily, one row per calendar day (~565 rows)

When exported flat, the rows desync after the first logging gap. To recover the
true calendar timeline, we treat:
  - row i of the Trend Weight column = trend value for (Oct 31, 2024 + i days)
  - row i of (Date, Scale) = the i-th logged day with that exact date

This script prints stats, writes a corrected CSV, and verifies that running
α=0.1 EWMA on a linearly-interpolated daily series reproduces MF's trend.
"""
from __future__ import annotations

import sys
from datetime import date, timedelta
from pathlib import Path

import openpyxl

XLSX = Path(sys.argv[1]) if len(sys.argv) > 1 else Path(
    r"C:\Users\drorm\Downloads\MacroFactor-20260518132238.xlsx.xlsx"
)

wb = openpyxl.load_workbook(XLSX, data_only=True)
ws = wb.active
print("Sheet:", ws.title, "rows:", ws.max_row, "cols:", ws.max_column)

header = [c.value for c in ws[1]]
print("Headers:", header)

# Find columns by header (handle variations)
def col_index(name_substring: str) -> int:
    for i, h in enumerate(header):
        if h and name_substring.lower() in str(h).lower():
            return i
    raise KeyError(name_substring)

ci_date = col_index("date")
ci_scale = col_index("weight (kg)")
ci_trend = col_index("trend")
print(f"Date col={ci_date}, Scale col={ci_scale}, Trend col={ci_trend}")

# Read all rows
rows = []
for r in ws.iter_rows(min_row=2, values_only=True):
    rows.append(r)
print(f"Total data rows: {len(rows)}")

# Build the two independent arrays
scale_log = []  # list of (date, weight) — only logged
trend_daily = []  # list of trend values — continuous from row 0 calendar day

for r in rows:
    d = r[ci_date]
    w = r[ci_scale]
    t = r[ci_trend]
    if d is not None and w is not None:
        # parse date
        if hasattr(d, "date"):
            d = d.date()
        elif isinstance(d, str):
            d = date.fromisoformat(d.split(" ")[0])
        scale_log.append((d, float(w)))
    if t is not None:
        trend_daily.append(float(t))

print(f"Logged scale entries: {len(scale_log)}")
print(f"Trend daily values: {len(trend_daily)}")

# Sort scale by date and confirm date range
scale_log.sort(key=lambda x: x[0])
first_logged = scale_log[0][0]
last_logged = scale_log[-1][0]
print(f"First logged: {first_logged}, last logged: {last_logged}")
total_days_logged_range = (last_logged - first_logged).days + 1
print(f"Calendar days in logged range: {total_days_logged_range}")
print(f"Trend daily length matches total days? {len(trend_daily) == total_days_logged_range}")

# So trend_daily[i] corresponds to calendar day (first_logged + i days)
trend_by_date = {}
for i, t in enumerate(trend_daily):
    d = first_logged + timedelta(days=i)
    trend_by_date[d] = t

print(f"\nFirst 5 trend values:")
for i in range(5):
    d = first_logged + timedelta(days=i)
    print(f"  {d}: {trend_daily[i]:.4f}")

print(f"\nLast 5 trend values:")
for i in range(max(0, len(trend_daily) - 5), len(trend_daily)):
    d = first_logged + timedelta(days=i)
    print(f"  {d}: {trend_daily[i]:.4f}")

# Build interpolated daily scale series
logged_by_date = {d: w for d, w in scale_log}
print(f"\nFirst few logged: {scale_log[:5]}")
print(f"Last few logged: {scale_log[-5:]}")

# Compute interpolated daily scale
daily_scale = []
for i in range(len(trend_daily)):
    d = first_logged + timedelta(days=i)
    if d in logged_by_date:
        daily_scale.append((d, logged_by_date[d], False))
        continue
    # find prev logged
    prev_d = None
    next_d = None
    for j in range(i - 1, -1, -1):
        dj = first_logged + timedelta(days=j)
        if dj in logged_by_date:
            prev_d = dj
            break
    for j in range(i + 1, len(trend_daily)):
        dj = first_logged + timedelta(days=j)
        if dj in logged_by_date:
            next_d = dj
            break
    if prev_d is not None and next_d is not None:
        span = (next_d - prev_d).days
        offset = (d - prev_d).days
        ratio = offset / span if span else 0
        interp = logged_by_date[prev_d] + (logged_by_date[next_d] - logged_by_date[prev_d]) * ratio
    elif prev_d is not None:
        interp = logged_by_date[prev_d]
    elif next_d is not None:
        interp = logged_by_date[next_d]
    else:
        interp = 0
    daily_scale.append((d, interp, True))

# Run α=0.1 EWMA on the daily scale
ALPHA = 0.1
ewma = []
prev = None
for d, s, _ in daily_scale:
    if prev is None:
        prev = s
    else:
        prev = ALPHA * s + (1 - ALPHA) * prev
    ewma.append((d, prev))

# Compare ewma vs trend_daily
diffs = []
for i, (d, our_t) in enumerate(ewma):
    mf_t = trend_daily[i]
    diffs.append(abs(our_t - mf_t))

mean_abs = sum(diffs) / len(diffs)
max_abs = max(diffs)
print(f"\n=== Verification of alpha=0.1 against MF trend (calendar-aligned) ===")
print(f"Days compared: {len(diffs)}")
print(f"Mean abs error: {mean_abs:.4f} kg")
print(f"Max abs error: {max_abs:.4f} kg")

# Where is the max error?
max_idx = diffs.index(max_abs)
print(f"Max error at day {max_idx}: {ewma[max_idx][0]} | ours={ewma[max_idx][1]:.4f} | mf={trend_daily[max_idx]:.4f}")

# Sample a few values
print(f"\nSample days (ours vs MF):")
for sample_i in [0, 1, 7, 14, 45, 100, 200, 365, 500, len(ewma) - 1]:
    if sample_i < len(ewma):
        d, t = ewma[sample_i]
        print(f"  day {sample_i:>3} {d}: ours={t:.4f} mf={trend_daily[sample_i]:.4f} Δ={t - trend_daily[sample_i]:+.4f}")

# Now compute KPIs for each window with MF's trend and various averaging strategies
print("\n=== MF KPIs from raw trend column (using calendar-aligned data) ===")
# today = last day of trend_daily
today = first_logged + timedelta(days=len(trend_daily) - 1)
print(f"Today (last trend day): {today}")
print(f"Today's trend: {trend_daily[-1]:.4f}")

windows = {
    "1W": 7,
    "1M": 30,
    "3M": 90,
    "6M": 180,
    "1Y": 365,
    "All": len(trend_daily),
}

# Build a list of (date, scale, trend)
calendar = [(daily_scale[i][0], daily_scale[i][1], trend_daily[i]) for i in range(len(trend_daily))]

def kpi(window_days: int, end_idx: int):
    """Compute KPIs for a window ending at end_idx with given # days back."""
    start_idx_a = max(0, end_idx - window_days + 1)  # inclusive: N days total
    start_idx_b = max(0, end_idx - window_days)       # one day earlier: N+1 days total
    out = {}
    for label, sidx in [("incl(N)", start_idx_a), ("incl(N+1)", start_idx_b)]:
        slice_ = calendar[sidx : end_idx + 1]
        n = len(slice_)
        avg_scale = sum(s for _, s, _ in slice_) / n
        avg_trend = sum(t for _, _, t in slice_) / n
        diff_trend = slice_[-1][2] - slice_[0][2]
        out[label] = (n, avg_scale, avg_trend, diff_trend)
    return out

end_idx = len(trend_daily) - 1
mf_targets = {
    "1W": (95.8, -0.5),
    "1M": (96.0, -0.7),
    "3M": (96.0, -0.7),  # user-reported; suspect mis-quote
    "6M": (97.1, -3.8),
    "1Y": (100.7, -13.9),
    "All": (107.0, -30.0),
}

def r1(x):
    """Round half-to-even at 1 decimal, like banker's rounding (Python default)."""
    return round(x, 1)

print("\n  >>> Compare two diff strategies: raw-then-round vs round-each-then-subtract")
for w_name, w_days in windows.items():
    if w_name == "All":
        w_days = len(trend_daily)
    print(f"\n  {w_name} (MF target: avg={mf_targets[w_name][0]} diff={mf_targets[w_name][1]}):")
    n_days = w_days
    sidx = max(0, end_idx - n_days + 1)
    slice_ = calendar[sidx : end_idx + 1]
    n = len(slice_)
    avg_t = sum(t for _, _, t in slice_) / n
    raw_diff = slice_[-1][2] - slice_[0][2]
    rounded_diff = r1(slice_[-1][2]) - r1(slice_[0][2])
    print(
        f"    n={n} avg={r1(avg_t)} | raw_diff={raw_diff:+.4f}->{r1(raw_diff):+.1f} | rounded_first={r1(slice_[0][2])} rounded_last={r1(slice_[-1][2])} rounded_diff={r1(rounded_diff):+.1f}"
    )

# For 1W especially, what's trend at boundary days?
print("\n--- Trend values near today ---")
for offset in range(15):
    i = end_idx - offset
    d = first_logged + timedelta(days=i)
    print(f"  day {i} ({d}, T-{offset}): trend={trend_daily[i]:.4f}")

# Save calendar-aligned CSV for downstream use
out_path = Path("scripts/macrofactor-calendar-aligned.csv")
out_path.parent.mkdir(exist_ok=True)
with out_path.open("w", encoding="utf-8") as f:
    f.write("date,logged_scale,interp_scale,mf_trend,our_trend\n")
    for i in range(len(trend_daily)):
        d, interp_s, was_interp = daily_scale[i]
        logged = logged_by_date.get(d, "")
        f.write(f"{d},{logged},{interp_s:.4f},{trend_daily[i]:.4f},{ewma[i][1]:.4f}\n")
print(f"\nWrote {out_path}")

# Save logged-only JSON for Playwright test seeding
import json
logged_json_path = Path("scripts/macrofactor-logged.json")
with logged_json_path.open("w", encoding="utf-8") as f:
    json.dump(
        {
            "first_logged_date": str(scale_log[0][0]),
            "last_logged_date": str(scale_log[-1][0]),
            "today": str(today),
            "expected_kpis": {
                "1W": {"avg": 95.8, "diff": -0.5},
                "1M": {"avg": 96.0, "diff": -0.7},
                "3M": {"avg": 96.0, "diff": -0.7},
                "6M": {"avg": 97.1, "diff": -3.8},
                "1Y": {"avg": 100.7, "diff": -13.9},
                "All": {"avg": 107.0, "diff": -30.0},
            },
            "computed_kpis_from_mf_trend": {
                "1W": {"avg": 95.8, "diff": -0.4},
                "1M": {"avg": 96.0, "diff": -0.7},
                "3M": {"avg": 96.5, "diff": -1.2},
                "6M": {"avg": 97.1, "diff": -3.7},
                "1Y": {"avg": 100.7, "diff": -13.9},
                "All": {"avg": 107.0, "diff": -29.9},
            },
            "entries": [
                {"date": str(d), "weight": w} for d, w in scale_log
            ],
        },
        f,
        indent=2,
    )
print(f"Wrote {logged_json_path} ({len(scale_log)} entries)")
